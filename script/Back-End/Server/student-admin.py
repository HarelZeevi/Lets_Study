import tkinter
import json
from tkinter.filedialog import askdirectory, askopenfilename

import openpyxl as xlsx
import requests as requests

ascii_txt = '''


░█████╗░██████╗░███╗░░░███╗██╗███╗░░██╗  ░██████╗██╗░░░██╗░██████╗████████╗███████╗███╗░░░███╗░░░██████╗░██╗░░░██╗
██╔══██╗██╔══██╗████╗░████║██║████╗░██║  ██╔════╝╚██╗░██╔╝██╔════╝╚══██╔══╝██╔════╝████╗░████║░░░██╔══██╗╚██╗░██╔╝
███████║██║░░██║██╔████╔██║██║██╔██╗██║  ╚█████╗░░╚████╔╝░╚█████╗░░░░██║░░░█████╗░░██╔████╔██║░░░██████╔╝░╚████╔╝░
██╔══██║██║░░██║██║╚██╔╝██║██║██║╚████║  ░╚═══██╗░░╚██╔╝░░░╚═══██╗░░░██║░░░██╔══╝░░██║╚██╔╝██║░░░██╔═══╝░░░╚██╔╝░░
██║░░██║██████╔╝██║░╚═╝░██║██║██║░╚███║  ██████╔╝░░░██║░░░██████╔╝░░░██║░░░███████╗██║░╚═╝░██║██╗██║░░░░░░░░██║░░░
╚═╝░░╚═╝╚═════╝░╚═╝░░░░░╚═╝╚═╝╚═╝░░╚══╝  ╚═════╝░░░░╚═╝░░░╚═════╝░░░░╚═╝░░░╚══════╝╚═╝░░░░░╚═╝╚═╝╚═╝░░░░░░░░╚═╝░░░

'''


# get folder using tkinter browse option
def get_folder():
    tkinter.Tk().withdraw()
    filename = askopenfilename(title="Choose Excel File",
                               initialdir='/')  # show an "Open" dialog box and return the path to the selected file
    return filename


def sign_in_admin(id, pswd):
    URL = "http://localhost:1234/api/admins/signIn"
    DATA = {'id': id, 'password': pswd}
    res = requests.post(url=URL, data=DATA)
    return res.json()['accessToken']

def add_student(DATA, JWT):
    URL = "http://localhost:1234/api/students/"
    HEADERS = {"authorization": "Bearer " + JWT}
    res = requests.post(url=URL, data=DATA, headers=HEADERS)
    return(res)

def change_xlsx(path, JWT):
    wrbk = xlsx.load_workbook(path)
    sh = wrbk.active
    students = []

    # iterate through excel and display data
    for i in range(2, sh.max_row + 1):
        student = {}
        # id
        id = sh.cell(row=i, column=1)
        student["id"] = id.value
        # grade
        grade = sh.cell(row=i, column=2)
        student["grade"] = grade.value
        # classnum
        classnum = sh.cell(row=i, column=3)
        student["classnum"] = classnum.value
        print(student)
        students.append(student)
        add_student(student, JWT)
    print(students)
    return students


def main():
    print(ascii_txt)
    print()
    a_id = "342665978"
    a_pswd = "1234"
    JWT = sign_in_admin(a_id, a_pswd)
    change_xlsx(get_folder(), JWT)


if __name__ == "__main__":
    main()
